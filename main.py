import os
import re
import math
import random
import asyncio
from io import BytesIO
from datetime import datetime, timedelta, date, time
from zoneinfo import ZoneInfo

import asyncpg
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from aiogram import Bot, Dispatcher, F
from aiogram.types import Message, BufferedInputFile
from aiogram.filters import Command, CommandStart

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
ADMIN_IDS_RAW = os.getenv("ADMIN_IDS", "").strip()
ADMIN_IDS = {int(x.strip()) for x in ADMIN_IDS_RAW.split(",") if x.strip().isdigit()} if ADMIN_IDS_RAW else set()

if not BOT_TOKEN or not DATABASE_URL:
    raise RuntimeError("Missing BOT_TOKEN or DATABASE_URL")

TZ = ZoneInfo("Asia/Colombo")

# ====== 写死班次（斯里兰卡）======
SHIFT_DAY_START = time(7, 0)
SHIFT_DAY_END = time(19, 0)   # 19:00 切到夜班（>=19:00 是夜班）

# ====== 写死规则（次数限制 + 单次上限分钟）======
BREAK_RULES = {
    "PEE":   {"minutes": 6,  "max_count": 3, "aliases": ["小便", "尿", "pee"]},
    "POOP":  {"minutes": 15, "max_count": 2, "aliases": ["大便", "拉屎", "poop"]},
    "EAT":   {"minutes": 30, "max_count": 3, "aliases": ["吃饭", "eat", "meal", "饭"]},
    "SMOKE": {"minutes": 10, "max_count": 5, "aliases": ["抽烟", "抽", "smoke"]},
}

WORKIN_ALIASES = ["上班", "开工", "开工了", "上工", "in", "start", "work in", "到岗", "开工吧", "开工呀"]
BACK_ALIASES = ["回", "回来", "back", "/back", "1", "/1", "结束", "return", "回来了"]

TOTAL_BREAK_LIMIT_MIN = 188

WORKIN_REPLIES = [
    "✅ 上班已记录，辛苦啦，今天稳住就赢了。",
    "✅ 到岗了，挺靠谱的，继续保持。",
    "✅ 已打卡上班，感谢配合。",
    "✅ 上班记录成功，今天加油。",
    "✅ 到岗确认，做得不错。",
]

def is_admin(user_id: int) -> bool:
    return (not ADMIN_IDS) or (user_id in ADMIN_IDS)

def norm_text(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^\w\u4e00-\u9fff\s]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def contains_any(text: str, aliases: list[str]) -> bool:
    for a in aliases:
        if a and a.lower() in text:
            return True
    return False

def get_shift(local_dt: datetime) -> tuple[date, str]:
    """
    返回：(shift_date, shift_type)
    - DAY: 07:00-18:59，shift_date = 当天日期
    - NIGHT: 19:00-06:59
        - 若时间在 19:00-23:59：shift_date = 当天日期
        - 若时间在 00:00-06:59：shift_date = 前一天日期（夜班属于前一天19点那班）
    """
    t = local_dt.time()
    d = local_dt.date()

    if SHIFT_DAY_START <= t < SHIFT_DAY_END:
        return d, "DAY"
    # NIGHT
    if t < SHIFT_DAY_START:
        return (d - timedelta(days=1)), "NIGHT"
    return d, "NIGHT"

def local_str(dt_utc: datetime | None) -> str:
    if not dt_utc:
        return ""
    if dt_utc.tzinfo is None:
        dt_utc = dt_utc.replace(tzinfo=ZoneInfo("UTC"))
    return dt_utc.astimezone(TZ).strftime("%Y-%m-%d %H:%M:%S")

def kind_cn(kind: str) -> str:
    return {"PEE": "小便", "POOP": "大便", "EAT": "吃饭", "SMOKE": "抽烟"}.get(kind, kind)

def compute_violation(counts: dict, mins: dict) -> tuple[bool, str]:
    reasons = []
    total_min = sum(mins.values())
    if total_min > TOTAL_BREAK_LIMIT_MIN:
        reasons.append(f"总离岗{total_min}>{TOTAL_BREAK_LIMIT_MIN}")

    for k, rule in BREAK_RULES.items():
        if counts.get(k, 0) > rule["max_count"]:
            reasons.append(f"{k}次数{counts[k]}>{rule['max_count']}")
    return (len(reasons) > 0, "; ".join(reasons))

DDL = """
CREATE TABLE IF NOT EXISTS tg_groups (
  chat_id BIGINT PRIMARY KEY,
  title TEXT,
  created_at TIMESTAMP NOT NULL DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS known_users (
  chat_id BIGINT NOT NULL,
  user_id BIGINT NOT NULL,
  user_name TEXT,
  first_seen TIMESTAMP NOT NULL DEFAULT NOW(),
  last_seen TIMESTAMP NOT NULL DEFAULT NOW(),
  PRIMARY KEY(chat_id, user_id)
);

CREATE TABLE IF NOT EXISTS shift_attendance (
  chat_id BIGINT NOT NULL,
  shift_date DATE NOT NULL,
  shift_type TEXT NOT NULL, -- DAY / NIGHT
  user_id BIGINT NOT NULL,
  user_name TEXT,
  first_in_at TIMESTAMP NULL,      -- 记录第一次“上班”
  last_action_at TIMESTAMP NULL,

  pee_count INT NOT NULL DEFAULT 0,
  poop_count INT NOT NULL DEFAULT 0,
  eat_count INT NOT NULL DEFAULT 0,
  smoke_count INT NOT NULL DEFAULT 0,

  pee_min INT NOT NULL DEFAULT 0,
  poop_min INT NOT NULL DEFAULT 0,
  eat_min INT NOT NULL DEFAULT 0,
  smoke_min INT NOT NULL DEFAULT 0,

  violation BOOLEAN NOT NULL DEFAULT FALSE,
  violation_reason TEXT NOT NULL DEFAULT '',

  PRIMARY KEY(chat_id, shift_date, shift_type, user_id)
);

CREATE INDEX IF NOT EXISTS idx_shift_date ON shift_attendance(chat_id, shift_date);

-- 进行中的离岗会话（重启不丢）
CREATE TABLE IF NOT EXISTS break_sessions (
  chat_id BIGINT NOT NULL,
  user_id BIGINT NOT NULL,
  shift_date DATE NOT NULL,
  shift_type TEXT NOT NULL,      -- DAY/NIGHT
  kind TEXT NOT NULL,            -- PEE/POOP/EAT/SMOKE
  start_at TIMESTAMP NOT NULL DEFAULT NOW(),
  limit_min INT NOT NULL,

  start_msg_id BIGINT NULL,      -- “已记录/请在xx前回来”
  remind_msg_id BIGINT NULL,     -- “已到上限请回来”

  PRIMARY KEY(chat_id, user_id, shift_date, shift_type)
);

CREATE INDEX IF NOT EXISTS idx_break_sessions_chat ON break_sessions(chat_id);
"""

pool: asyncpg.Pool | None = None

async def db_init():
    global pool
    last_err = None
    for i in range(30):
        try:
            pool = await asyncpg.create_pool(DATABASE_URL, min_size=1, max_size=5)
            async with pool.acquire() as conn:
                await conn.execute("SELECT 1;")
                await conn.execute(DDL)
            print("[db] connected & migrated")
            return
        except Exception as e:
            last_err = e
            print(f"[db] connect failed ({i+1}/30): {e}")
            await asyncio.sleep(2)
    raise RuntimeError(f"DB connect failed after retries: {last_err}")

async def upsert_group(chat_id: int, title: str):
    async with pool.acquire() as conn:
        await conn.execute(
            "INSERT INTO tg_groups(chat_id, title) VALUES($1,$2) "
            "ON CONFLICT(chat_id) DO UPDATE SET title=EXCLUDED.title",
            chat_id, title
        )

async def upsert_known_user(chat_id: int, user_id: int, user_name: str):
    async with pool.acquire() as conn:
        await conn.execute(
            "INSERT INTO known_users(chat_id, user_id, user_name) VALUES($1,$2,$3) "
            "ON CONFLICT(chat_id, user_id) DO UPDATE SET user_name=EXCLUDED.user_name, last_seen=NOW()",
            chat_id, user_id, user_name
        )

async def get_or_create_shift_row(chat_id: int, shift_date: date, shift_type: str, user_id: int, user_name: str):
    async with pool.acquire() as conn:
        await conn.execute(
            "INSERT INTO shift_attendance(chat_id, shift_date, shift_type, user_id, user_name) "
            "VALUES($1,$2,$3,$4,$5) "
            "ON CONFLICT(chat_id, shift_date, shift_type, user_id) DO UPDATE SET user_name=EXCLUDED.user_name",
            chat_id, shift_date, shift_type, user_id, user_name
        )

async def update_workin(chat_id: int, shift_date: date, shift_type: str, user_id: int):
    async with pool.acquire() as conn:
        await conn.execute(
            """
            UPDATE shift_attendance
            SET first_in_at = COALESCE(first_in_at, NOW()),
                last_action_at = NOW()
            WHERE chat_id=$1 AND shift_date=$2 AND shift_type=$3 AND user_id=$4
            """,
            chat_id, shift_date, shift_type, user_id
        )

async def get_active_session(chat_id: int, shift_date: date, shift_type: str, user_id: int):
    async with pool.acquire() as conn:
        return await conn.fetchrow(
            """
            SELECT kind, start_at, limit_min, start_msg_id, remind_msg_id
            FROM break_sessions
            WHERE chat_id=$1 AND shift_date=$2 AND shift_type=$3 AND user_id=$4
            """,
            chat_id, shift_date, shift_type, user_id
        )

async def start_break_session(chat_id: int, shift_date: date, shift_type: str, user_id: int, kind: str, limit_min: int, start_msg_id: int):
    async with pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO break_sessions(chat_id, user_id, shift_date, shift_type, kind, limit_min, start_msg_id)
            VALUES($1,$2,$3,$4,$5,$6,$7)
            ON CONFLICT(chat_id, user_id, shift_date, shift_type)
            DO UPDATE SET kind=EXCLUDED.kind, start_at=NOW(), limit_min=EXCLUDED.limit_min,
                         start_msg_id=EXCLUDED.start_msg_id, remind_msg_id=NULL
            """,
            chat_id, user_id, shift_date, shift_type, kind, limit_min, start_msg_id
        )

async def set_remind_msg(chat_id: int, shift_date: date, shift_type: str, user_id: int, remind_msg_id: int):
    async with pool.acquire() as conn:
        await conn.execute(
            """
            UPDATE break_sessions SET remind_msg_id=$5
            WHERE chat_id=$1 AND shift_date=$2 AND shift_type=$3 AND user_id=$4
            """,
            chat_id, shift_date, shift_type, user_id, remind_msg_id
        )

async def end_break_session(chat_id: int, shift_date: date, shift_type: str, user_id: int):
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT kind, start_at, limit_min, start_msg_id, remind_msg_id
            FROM break_sessions
            WHERE chat_id=$1 AND shift_date=$2 AND shift_type=$3 AND user_id=$4
            """,
            chat_id, shift_date, shift_type, user_id
        )
        if not row:
            return None
        await conn.execute(
            """
            DELETE FROM break_sessions
            WHERE chat_id=$1 AND shift_date=$2 AND shift_type=$3 AND user_id=$4
            """,
            chat_id, shift_date, shift_type, user_id
        )
        return row

async def add_break_result(chat_id: int, shift_date: date, shift_type: str, user_id: int, kind: str, used_seconds: int):
    used_min = max(1, math.ceil(used_seconds / 60))

    async with pool.acquire() as conn:
        # 次数+1，用时累加
        await conn.execute(
            f"""
            UPDATE shift_attendance
            SET {kind.lower()}_count = {kind.lower()}_count + 1,
                {kind.lower()}_min   = {kind.lower()}_min + $5,
                last_action_at = NOW()
            WHERE chat_id=$1 AND shift_date=$2 AND shift_type=$3 AND user_id=$4
            """,
            chat_id, shift_date, shift_type, user_id, used_min
        )

        row = await conn.fetchrow(
            """
            SELECT pee_count, poop_count, eat_count, smoke_count,
                   pee_min, poop_min, eat_min, smoke_min
            FROM shift_attendance
            WHERE chat_id=$1 AND shift_date=$2 AND shift_type=$3 AND user_id=$4
            """,
            chat_id, shift_date, shift_type, user_id
        )

        counts = {"PEE": row["pee_count"], "POOP": row["poop_count"], "EAT": row["eat_count"], "SMOKE": row["smoke_count"]}
        mins = {"PEE": row["pee_min"], "POOP": row["poop_min"], "EAT": row["eat_min"], "SMOKE": row["smoke_min"]}
        viol, reason = compute_violation(counts, mins)

        await conn.execute(
            """
            UPDATE shift_attendance
            SET violation=$5, violation_reason=$6
            WHERE chat_id=$1 AND shift_date=$2 AND shift_type=$3 AND user_id=$4
            """,
            chat_id, shift_date, shift_type, user_id, viol, reason
        )

    return used_min, counts, mins, viol, reason

async def export_xlsx(chat_id: int, start_d: date, end_d: date) -> bytes:
    async with pool.acquire() as conn:
        users = await conn.fetch(
            "SELECT user_id, user_name FROM known_users WHERE chat_id=$1 ORDER BY user_name NULLS LAST, user_id",
            chat_id
        )

        shifts = []
        d = start_d
        while d <= end_d:
            shifts.append((d, "DAY"))
            shifts.append((d, "NIGHT"))
            d += timedelta(days=1)

        rows = await conn.fetch(
            """
            SELECT shift_date, shift_type, user_id, user_name, first_in_at,
                   pee_count, poop_count, eat_count, smoke_count,
                   pee_min, poop_min, eat_min, smoke_min,
                   violation, violation_reason
            FROM shift_attendance
            WHERE chat_id=$1 AND shift_date >= $2 AND shift_date <= $3
            """,
            chat_id, start_d, end_d
        )

    data_map = {(r["shift_date"], r["shift_type"], int(r["user_id"])): r for r in rows}

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"

    headers = [
        "shift_date", "shift_type", "user_id", "user_name",
        "clock_in_time(Colombo)", "status",
        "break_total_min",
        "eat_count", "smoke_count", "pee_count", "poop_count",
        "eat_min", "smoke_min", "pee_min", "poop_min",
        "violation", "violation_reason"
    ]
    ws.append(headers)
    bold = Font(bold=True)
    for c in ws[1]:
        c.font = bold

    red_fill = PatternFill("solid", fgColor="FFCCCC")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")

    for (sd, st) in shifts:
        for u in users:
            uid = int(u["user_id"])
            uname = u["user_name"] or str(uid)
            r = data_map.get((sd, st, uid))

            if r and r["first_in_at"]:
                status = "OK"
                cin = local_str(r["first_in_at"])
            else:
                status = "MISSING"
                cin = ""

            if r:
                eat_c, smk_c, pee_c, pop_c = r["eat_count"], r["smoke_count"], r["pee_count"], r["poop_count"]
                eat_m, smk_m, pee_m, pop_m = r["eat_min"], r["smoke_min"], r["pee_min"], r["poop_min"]
                viol, reason = r["violation"], r["violation_reason"]
            else:
                eat_c = smk_c = pee_c = pop_c = 0
                eat_m = smk_m = pee_m = pop_m = 0
                viol, reason = False, ""

            total_break = eat_m + smk_m + pee_m + pop_m

            ws.append([
                sd.strftime("%Y-%m-%d"), st, uid, uname,
                cin, status,
                total_break,
                eat_c, smk_c, pee_c, pop_c,
                eat_m, smk_m, pee_m, pop_m,
                "YES" if viol else "NO",
                reason
            ])

            row_idx = ws.max_row
            if status == "MISSING":
                for cell in ws[row_idx]:
                    cell.fill = red_fill
            elif viol:
                for cell in ws[row_idx]:
                    cell.fill = yellow_fill

    ws2 = wb.create_sheet("rules")
    ws2.append(["timezone", "Asia/Colombo"])
    ws2.append(["shift DAY", "07:00-18:59"])
    ws2.append(["shift NIGHT", "19:00-06:59 (belongs to shift_date of 19:00 day)"])
    ws2.append(["TOTAL_BREAK_LIMIT_MIN", TOTAL_BREAK_LIMIT_MIN])
    ws2.append([])
    ws2.append(["TYPE", "minutes_each(单次上限)", "max_count(次数上限)", "aliases"])
    for k, v in BREAK_RULES.items():
        ws2.append([k, v["minutes"], v["max_count"], ", ".join(v["aliases"])])

    for sheet in (ws, ws2):
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            sheet.column_dimensions[col_letter].width = min(max_len + 2, 40)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ===== Bot =====
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

@dp.message(CommandStart())
async def start(m: Message):
    if m.chat.type == "private":
        if not is_admin(m.from_user.id):
            await m.reply("无权限。")
            return
        await m.reply(
            "✅ 打卡机器人已启动（写死规则版）\n\n"
            "群里发：上班/in/开工；离岗：吃饭/eat、小便/pee、大便/poop、抽烟/抽/smoke\n"
            "回来：回 / back / 1 / 结束\n"
            "（不打下班，按斯里兰卡时间自动分白班/夜班）\n\n"
            "导出：/export <chat_id> YYYY-MM-DD YYYY-MM-DD\n"
            "示例：/export -1001234567890 2026-02-01 2026-02-05\n\n"
            "⚠️ 提醒：机器人需要群管理员权限才能删提示消息。"
        )
    else:
        await upsert_group(m.chat.id, m.chat.title or "")
        await m.reply("✅ 已加入群。直接发 上班/吃饭/eat/抽烟/小便/大便 即可记录；回来发：回/back/1/结束。")

@dp.message(Command("export"))
async def export_cmd(m: Message):
    if m.chat.type != "private":
        return
    if not is_admin(m.from_user.id):
        await m.reply("无权限。")
        return

    parts = m.text.split()
    if len(parts) not in (2, 4):
        await m.reply("用法：/export <chat_id> 或 /export <chat_id> YYYY-MM-DD YYYY-MM-DD")
        return

    try:
        chat_id = int(parts[1])
    except:
        await m.reply("chat_id 格式不对，例如 -1001234567890")
        return

    if len(parts) == 2:
        start_d = date.today()
        end_d = start_d
    else:
        try:
            start_d = datetime.strptime(parts[2], "%Y-%m-%d").date()
            end_d = datetime.strptime(parts[3], "%Y-%m-%d").date()
        except:
            await m.reply("日期格式不对：YYYY-MM-DD")
            return

    data = await export_xlsx(chat_id, start_d, end_d)
    fn = f"attendance_{chat_id}_{start_d.strftime('%Y%m%d')}_{end_d.strftime('%Y%m%d')}.xlsx"
    await m.reply_document(BufferedInputFile(data, filename=fn), caption="✅ 导出完成（红色=MISSING，黄色=违规）")

@dp.message(F.chat.type.in_(["group", "supergroup"]) & Command("chatid"))
async def chatid_cmd(m: Message):
    await m.reply(f"chat_id={m.chat.id}\n群名：{m.chat.title or ''}")

@dp.message(F.chat.type.in_(["group", "supergroup"]) & F.text)
async def group_listener(m: Message):
    await upsert_group(m.chat.id, m.chat.title or "")

    text = norm_text(m.text)
    if not text:
        return

    uid = m.from_user.id
    uname = m.from_user.full_name or (m.from_user.username or str(uid))
    await upsert_known_user(m.chat.id, uid, uname)

    now_local = datetime.now(tz=TZ)
    shift_date, shift_type = get_shift(now_local)
    await get_or_create_shift_row(m.chat.id, shift_date, shift_type, uid, uname)

    # ① 回来：结束离岗
    if contains_any(text, BACK_ALIASES):
        sess = await end_break_session(m.chat.id, shift_date, shift_type, uid)
        if not sess:
            await m.reply("你当前没有进行中的离岗记录。")
            return

        kind = sess["kind"]
        start_at = sess["start_at"]
        limit_min = int(sess["limit_min"])
        start_msg_id = sess["start_msg_id"]
        remind_msg_id = sess["remind_msg_id"]

        # 删除“开始提示 + 超时提醒”（需要机器人有删消息权限）
        for mid in [start_msg_id, remind_msg_id]:
            if mid:
                try:
                    await bot.delete_message(m.chat.id, int(mid))
                except:
                    pass

        # 计算真实用时
        # asyncpg 取出的 timestamp 通常是 naive，这里按 Colombo 解释
        if start_at.tzinfo is None:
            start_local = start_at.replace(tzinfo=TZ)
        else:
            start_local = start_at.astimezone(TZ)

        used_seconds = int((datetime.now(tz=TZ) - start_local).total_seconds())
        used_min, counts, mins, viol, reason = await add_break_result(m.chat.id, shift_date, shift_type, uid, kind, used_seconds)

        overtime = used_min > limit_min
        left_times = BREAK_RULES[kind]["max_count"] - counts[kind]
        total_break = sum(mins.values())

        msg = (
            f"✅ {uname} {kind_cn(kind)} 本次结束，用时 {used_min} 分钟（上限 {limit_min} 分钟）{'⚠️已超时' if overtime else ''}\n"
            f"本班累计：{kind_cn(kind)} 第 {counts[kind]} 次（剩余 {max(left_times,0)} 次），累计离岗 {total_break} 分钟。\n"
            f"{'⚠️违规：'+reason if viol else '正常'}"
        )
        await m.reply(msg)
        return

    # ② 上班：回复欣慰话
    if contains_any(text, WORKIN_ALIASES):
        await update_workin(m.chat.id, shift_date, shift_type, uid)
        await m.reply(random.choice(WORKIN_REPLIES))
        return

    # ③ 离岗开始
    for kind, rule in BREAK_RULES.items():
        if contains_any(text, rule["aliases"]):
            active = await get_active_session(m.chat.id, shift_date, shift_type, uid)
            if active:
                await m.reply(f"你正在 {kind_cn(active['kind'])} 中，请先发送：回 / back / 1 / 结束。")
                return

            # 当前次数
            async with pool.acquire() as conn:
                r = await conn.fetchrow(
                    f"""
                    SELECT {kind.lower()}_count AS c
                    FROM shift_attendance
                    WHERE chat_id=$1 AND shift_date=$2 AND shift_type=$3 AND user_id=$4
                    """,
                    m.chat.id, shift_date, shift_type, uid
                )
            used_count = int(r["c"]) if r else 0
            next_count = used_count + 1
            max_count = rule["max_count"]
            limit_min = rule["minutes"]

            deadline = datetime.now(tz=TZ) + timedelta(minutes=limit_min)
            start_msg = await m.reply(
                f"⏰ ✅ 已记录：{uname} {kind_cn(kind)}（第 {next_count} 次 / 限制 {max_count} 次）\n"
                f"请在 {deadline.strftime('%H:%M')} 前回来：回 / back / 1 / 结束"
            )
            await start_break_session(m.chat.id, shift_date, shift_type, uid, kind, limit_min, start_msg.message_id)

            # 到点提醒（简单版：重启会丢提醒，但“回来总结/累计”不会丢）
            async def remind_later(chat_id: int, user_id: int, sd: date, st: str, kind_: str, limit_: int):
                await asyncio.sleep(limit_ * 60)
                still = await get_active_session(chat_id, sd, st, user_id)
                if still and still["kind"] == kind_:
                    try:
                        rm = await bot.send_message(
                            chat_id,
                            f"⏰ @{m.from_user.username or uname} 的 {kind_cn(kind_)} 已到上限 {limit_} 分，请尽快回来：回 / back / 1 / 结束"
                        )
                        await set_remind_msg(chat_id, sd, st, user_id, rm.message_id)
                    except:
                        pass

            asyncio.create_task(remind_later(m.chat.id, uid, shift_date, shift_type, kind, limit_min))
            return

async def main():
    await db_init()
    print("[bot] polling started")
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
